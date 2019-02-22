'''
Rules to follow:
 - Teachers with only periods in {1, 2, 3, 4} are part-time on blue days, so all classes will have
   allowed periods of 1, 2, 3, 4. Same for {5, 6 ,7} and red days.
 - Semester courses that end up in the same period will be seen as constrained to be in the same period
 - Semester courses that are sem 1 will be seen as constrained to be in sem 1. Same for sem 2.
 - Homerooms will be ignored.
 - Room will be included as a constraint
 - Arabic key MUST be used to create list
    - Arabics are teamed

 - Teaming:
    - team_1: Sections teamed to have same students and different periods
    - team_2: Sections teamed to have different students and same period
        - Orchestra
        - team_period
    - team_3: Sections teamed to have the same students and same period, provided that students have requested class
        - Arabics
        - team_period_students
'''

from openpyxl import load_workbook
import re
import random

secList = load_workbook(filename='../../runs/constraint_files/SecList 8-14@1300.xlsx')
secSheet = secList['D1BC6ACA-A78F-4688-9C6D-42D84ED']

# Sections is a dict of sections coded as follows: {secID: [(0) courseID, (1) term, (2) course title, (3) teacherID, (4) room, (5) max, (6) [team1], (7) [team2], (8) [team3], (9) [allowedPeriods], (10) teacher first name, (11) teacher last name}
sections = {}
# {teacher's last name: [[(0) secID, (1) period, (2) term, (3) courseID, (4) courseTitle, (5) room, (6) cap, (7) teacher first name, (8) teacherID]]}
# NOTE: term notated as follows: 0 = year, 1 = semester 1, 2 = semester 2
teacherToSects = {}

teamings = load_workbook(filename='../../runs/constraint_files/Arabic and LS to Semesters.xlsx')
arabics = teamings['Sheet1']
learningSems = teamings['Sheet2']

humsAndIbets = load_workbook(filename='../../runs/constraint_files/17-18 HUMs at a glance.xlsx')
humSheet = humsAndIbets['Sheet3']
ibetSheet = humsAndIbets['IBETs']

# dict of year-long name of class to the two semester classes it translates to
teamedToSems = {}
# dict of semester classes to the year-long classes they're a part of.
semsToTeam = {}

# Read in data from teaming workbook (with Arabics and Learning Seminars)
for r in range(3, 28):
    curTeam = arabics.cell(row=r, column=1).value
    curSem1 = arabics.cell(row=r, column=10).value
    curSem2 = arabics.cell(row=r, column=21).value
    teamedToSems[curTeam] = {curSem1, curSem2}
    semsToTeam[curSem1] = curTeam
    semsToTeam[curSem2] = curTeam
for r in range(2, 26):
    curTeam = learningSems.cell(row=r, column=1).value
    curSem1 = learningSems.cell(row=r, column=10).value
    curSem2 = learningSems.cell(row=r, column=21).value
    teamedToSems[curTeam] = [curSem1, curSem2]
    semsToTeam[curSem1] = curTeam
    semsToTeam[curSem2] = curTeam

# Dict of section IDs to section IDs of clases with which it is teamed (team_1) in HUMs or CHUMs
secToHum = {}
# Dict of section IDs to section IDs of classes with which it is teamed (team_2) in HUMS only (IBETS still use limited periods)
humsTeam2 = {}
# Set of IBET course IDs
ibetIDs = {'4310T1', '984060', '1130T1'}

# Read in data from HUMs and IBETs workbook
n = 1
for r in range(1,45):
    sec1 = humSheet.cell(row=r, column=1).value
    sec2 = humSheet.cell(row=r, column=5).value
    sec3 = humSheet.cell(row=r, column=9).value
    secToHum[sec1] = [sec2]
    secToHum[sec2] = [sec1]
    if sec3:
        secToHum[sec1].append(sec3)
        secToHum[sec2].append(sec3)
        secToHum[sec3] = [sec1, sec2]
    if not humSheet.cell(row=r, column=4).value:
        if n:
            humsTeam2[sec1] = humSheet.cell(row=r+1, column=5).value
            humsTeam2[humSheet.cell(row=r+1, column=5).value] = sec1
            humsTeam2[sec2] = humSheet.cell(row=r+1, column=1).value
            humsTeam2[humSheet.cell(row=r+1, column=1).value] = sec2
        n = 1 - n
    else:
        n = 1

# Read in data from sections workbook
for r in range(2, 859):
    if secSheet.cell(row=r, column=2).value == 8:
        continue
    if secSheet.cell(row=r, column=6).value:
        teacher = secSheet.cell(row=r, column=6).value.split(", ")[0]
    elif secSheet.cell(row=r, column=5).value == "See Counselor":
        teacher = "Counselor"
    elif re.search(r'Health & PE', secSheet.cell(row=r, column=5).value):
        teacher = "Gym"
    else:
        teacher = "None"
    # Will post-process on "Gym." "None" and "Counselor" mean the teacher field should be removed, but counselor designation may be helpful later.
    if not teacher in teacherToSects.keys():
        teacherToSects[teacher] = []
    teacherToSects[teacher].append([])
    teacherToSects[teacher][-1].append(secSheet.cell(row=r, column=1).value)
    if secSheet.cell(row=r, column=2).value <= 7:
        teacherToSects[teacher][-1].append(secSheet.cell(row=r, column = 2).value)
    else:
        teacherToSects[teacher][-1].append(1)
    term = secSheet.cell(row=r, column=3).value
    if term == "YR":
        term = 0
    elif term == "S1":
        term = 1
    elif term == "S2":
        term = 2
    teacherToSects[teacher][-1].append(term)
    teacherToSects[teacher][-1].append(secSheet.cell(row=r, column =4).value.__str__())
    teacherToSects[teacher][-1].append(secSheet.cell(row=r, column=5).value)
    teacherToSects[teacher][-1].append(secSheet.cell(row=r, column=7).value)
    teacherToSects[teacher][-1].append(secSheet.cell(row=r, column=9).value)
    if teacher == "Counselor" or teacher == "Gym" or teacher == "None":
        teacherToSects[teacher][-1].append(None)
        teacherToSects[teacher][-1].append(None)
    else:
        teacherToSects[teacher][-1].append(secSheet.cell(row=r, column=6).value.split(", ")[1])
        # print(teacherToSects[teacher][-1])
        '''if teacher == "Jirari Scavotto":
            teacherToSects[teacher][-1].append("AJirari")'''
        #else:
        teacherToSects[teacher][-1].append(teacherToSects[teacher][-1][7][0] + teacher[:7].replace(" ", "").replace("-", ""))

# Parse data from sections workbook on teacher-by-teacher basis
for teacher in teacherToSects.keys():
    if teacher == "Gym":
        # Creates dict of gym teachers and how many gym classes they can teach
        teacherToNumber = {"Stile": 5, "Smith":5, "Potoker": 5, "Arthur": 6, "James": 5}
        # Creates dict of gym teachers to full teacher naming data
        teacherToName = {"Stile": ["Melissa", "Stile", "mmstile"], "Smith": ["Heidi", "Smith", "hesmith"], "Potoker": ["Barry", "Potoker", "bnpotoker"], "Arthur": ["David", "Arthur", "drarthur"], "James": ["Jeffery", "James", "jcjames"]}
        for curPer in teacherToSects[teacher]:
            gymTeach = random.choice([*teacherToNumber.keys()])
            while teacherToNumber[gymTeach] <= 0:
                gymTeach = random.choice([*teacherToNumber.keys()])
            teacherToNumber[gymTeach] -= 1
            sections[curPer[0]] = [curPer[3], curPer[2], curPer[4], teacherToName[gymTeach][2], curPer[5], curPer[6], None, None, None, None, teacherToName[gymTeach][0], gymTeach]
        continue
    if teacher == "Counselor":
        secID = 1
        for per in range(1,8):
            for sem in range(3):
                if sem == 0:
                    id = "000932"
                else:
                    id = "000933"
                sections["000932-" + "{:02d}".format(secID)] = [id, sem, "See Counselor", None, None, 10000, None, None, None, [per], None, None]
                secID += 1
        continue
    if teacher == "None":
        for curPer in teacherToSects[teacher]:
            sections[curPer[0]] = [curPer[3], curPer[2], curPer[4], curPer[8], curPer[5], curPer[6], None, None, None, [1,2,3,4,5,6,7], curPer[7], None]
        continue
    # perDist is a dict of {period: [indeces of classes in teacherToSects[teacher] during that period]}
    perDist = {l:[] for l in range(1,8)}
    perSet = set()
    for x in range(len(teacherToSects[teacher])):
        perDist[teacherToSects[teacher][x][1]].append(x)
        perSet.add(teacherToSects[teacher][x][1])
    # Remember to restrict allowed periods for IBETs and CHUMs outside of the restrictions listed here
    if perSet.issubset({1,2,3,4}):
        allowedPers = [1,2,3,4]
    elif perSet.issubset({5,6,7}):
        allowedPers = [5,6,7]
    else:
        allowedPers = [1,2,3,4,5,6,7]
    '''if teacher == "Waters":
        print(allowedPers)
        print(perDist)'''
    for per in range(1, 1+len(perDist.keys())):
        curPers = [teacherToSects[teacher][perDist[per][k]] for k in range(len(perDist[per]))]
        '''if teacher == "Waters":
            print(per)
            print(curPers)
            print("\n")'''
        # Handles periods during which a teacher is only teaching one course
        if len(curPers) == 1:
            curPer = curPers[0]
            sections[curPer[0]] = [curPer[3], curPer[2], curPer[4], curPer[8], curPer[5], curPer[6], None, None, None, allowedPers, curPer[7], teacher]
            # Deals with teamed classes (team_1)
            if curPer[0] in secToHum:
                sections[curPer[0]][6] = secToHum[curPer[0]]
                # Handles team_2 for HUMs
                if curPer[0] in humsTeam2:
                    sections[curPer[0]][7] = [humsTeam2[curPer[0]]]
                # Handles allowed periods for IBETs and CHUMs
                else:
                    if per in {5,6,7}:
                        sections[curPer[0]][9] = [5,6,7]
                    else:
                        if sections[curPer[0]][0] in ibetIDs:
                            sections[curPer[0]][9] = [1,2,3]
                        else:
                            sections[curPer[0]][9] = [1,2,3,4]
        else:
            # Builds set of sections in period still to handle
            secsToHandle = {curPers[k][0]:curPers[k] for k in range(len(curPers))}
            # Creates team_3 sections for arabics and learning seminars
            while set(secsToHandle.keys()).intersection(teamedToSems.keys()):
                year = [*set(secsToHandle.keys()).intersection(teamedToSems.keys())][0]
                for sec in teamedToSems[year]:
                    if sec in secsToHandle.keys():
                        curPer = secsToHandle[sec]
                        sections[sec] = [curPer[3], curPer[2], curPer[4], curPer[8], curPer[5], secsToHandle[year][6], None, None, [*teamedToSems[year].difference({sec})], allowedPers, curPer[7], teacher]
                        del secsToHandle[sec]
                del secsToHandle[year]
            # Loops over sections left to handle that aren't learning sems or arabics
            tempSecSet = set(secsToHandle.keys())
            for sec in tempSecSet:
                # Creates curPer, a variable storing the section data for sec.
                curPer = secsToHandle[sec]
                # Creates a section for sec
                sections[sec] = [curPer[3], curPer[2], curPer[4], curPer[8], curPer[5], curPer[6], None, None, None, allowedPers, curPer[7], teacher]
                if teacher == "FCPS Online Campus":
                    sections[sec][9] = [curPer[1]]
                # Deal with team_2 here (all classes that are the same period and not semesters 1 and 2 should be team_2. Meaning same period, same term; and same period, year and semester.). Otherwise, no teaming.
                # Builds term distribution for sections remaining in secsToHandle
                termDist = {0:[], 1:[], 2:[]}
                for tempSec in secsToHandle.values():
                    if tempSec[0] == sec:
                        continue
                    termDist[tempSec[2]].append(tempSec[0])
                if curPer[2] == 0:
                    sections[sec][7] = termDist[0] + termDist[1] + termDist[2]
                else:
                    sections[sec][7] = termDist[0] + termDist[curPer[2]]
                del secsToHandle[sec]

# Sections is a dict of sections coded as follows: {secID: [(0) courseID, (1) term, (2) course title, (3) teacherID, (4) room, (5) max, (6) [team1], (7) [team2], (8) [team3], (9) [allowedPeriods], (10) teacher first name, (11) teacher last name}
# teacherToSects = {teacher's last name: [[(0) secID, (1) period, (2) term, (3) courseID, (4) courseTitle, (5) room, (6) cap, (7) teacher first name, (8) teacherID]]}

# Writes sections and teachers to files
teacherFile = open("../../runs/constraint_files/full_school_teachers.txt", "w")
secFile = open("../../runs/constraint_files/full_school_sections.txt", "w")
courseFile = open("../../runs/constraint_files/full_school_courses.txt", "w")

teacherIDSet = set()
courseIDSet = set()
for section in sections.items():
    secFile.write(section[0] + "\n")
    if section[1][3]:
        secFile.write("teacher: " + section[1][3] + "\n")
    secFile.write("courseID: " + section[1][0] + "\n")
    if section[1][4]:
        secFile.write("room: " + section[1][4].__str__() + "\n")
    secFile.write("semester: " + section[1][1].__str__() + "\n")
    secFile.write("maxstudents: " + section[1][5].__str__() + "\n")
    if section[1][6]:
        for sec in section[1][6]:
            secFile.write("team_1: " + sec + "\n")
    if section[1][7]:
        for sec in section[1][7]:
            secFile.write("team_2: " + sec + "\n")
    if section[1][8]:
        for sec in section[1][8]:
            secFile.write("team_3: " + sec + "\n")
    if section[1][9]:
        secFile.write("allowed_periods: ")
        secFile.write(section[1][9][0].__str__())
        for k in range(1,len(section[1][9])):
            secFile.write("," + section[1][9][k].__str__())
        secFile.write("\n")

    if section[1][3] and not section[1][3] in teacherIDSet:
        teacherFile.write(section[1][10] + ", " + section[1][11] + ", " + section[1][3] + "\n")
        teacherIDSet.add(section[1][3])

    if section[1][0] and not section[1][0] in courseIDSet:
        courseFile.write("- | " + section[1][2].__str__() + "| " + section[1][0].__str__() + "| ")
        if section[1][1] == 0:
            courseFile.write("year\n")
        else:
            courseFile.write("semester\n")
        courseIDSet.add(section[1][0])
    # Below is the last line for writing secFile
    secFile.write("\n")