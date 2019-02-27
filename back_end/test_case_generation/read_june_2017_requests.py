from openpyxl import load_workbook
import re

# Handle 280292 (semester) to 280493 (year).
raise Exception("Didn't implement above change")

wb = load_workbook(filename='../../runs/constraint_files/6-1-17@1430.xlsx')
ws = wb['Sheet1']

file = open("../../runs/constraint_files/full_school_2017_students.txt", "w")

'''
curStud = ws.cell(row=2,column=1).value
# ADJUST THE AUTOMATIC 7 COURSES TO BE INTELLIGENT!
file.write(curStud.__str__() + ", Student, " + curStud.__str__() + ", " + ws.cell(row=2, column=3).value.__str__() + ", 7\n")
for r in range(2,16000):
    if not ws.cell(row=r,column=1).value:
        break
    if not ws.cell(row=r, column=1).value == curStud:
        curStud = ws.cell(row=r, column=1).value
        file.write("\n" + curStud.__str__() + ", Student, " + curStud.__str__() + ", " + ws.cell(row=r, column=3).value.__str__() + ", 7\n")
    if not ws.cell(row=r, column=5).value == "See Counselor":
        file.write(ws.cell(row=r,column=4).value.__str__()+"\n")
'''

#'''
curStud = 1
r = 2
# FINISH MODIFYING to count how many course requests each student has before printing to text file (store course requests in list)
while ws.cell(row=r, column=1).value:
    studRequests = []
    while ws.cell(row=r,column=1).value==curStud:
        #print("HI")
        if ws.cell(row=r,column=5).value == "See Counselor":
            r+=1
            continue
        sectID = ws.cell(row=r, column=4).value.__str__()
        if sectID == "503000":
            studRequests.append("3190T1")
            studRequests.append("313753")
        elif sectID == "505000":
            studRequests.append("313754")
            studRequests.append("316055")
        elif sectID == "501000":
            studRequests.append("3190T1")
            studRequests.append("314351")
        elif sectID == "502000":
            studRequests.append("313753")
            studRequests.append("313754")
        elif sectID == "504000":
            studRequests.append("3190T1")
            studRequests.append("313754")
        elif sectID == "231904":
            studRequests.append("231905")
        elif sectID == "2340T1":
            studRequests.append("234093")
        else:
            studRequests.append(ws.cell(row=r, column=4).value.__str__())
        r += 1
    file.write(curStud.__str__() + ", Student, " + curStud.__str__() + ", " + ws.cell(row=r-1, column=3).value.__str__() + ", " + len(studRequests).__str__() + "\n")
    for x in range(len(studRequests)):
        file.write(studRequests[x] + "\n")
    file.write("\n")
    #print(curStud)
    #print(studRequests)
    curStud = ws.cell(row=r, column=1).value
#'''
