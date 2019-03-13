# Script used to calculate the number of unfulfillable requests in the 2017-2018 requests data
# Spreadsheets needed: 6-1-17@1430.xlsx and SecList 8-14@1300.xlsx

from openpyxl import load_workbook
import re

stud_wb = load_workbook(filename="../../runs/constraint_files/6-1-17@1430.xlsx")
sec_wb = load_workbook(filename="../../runs/constraint_files/SecList 8-14@1300.xlsx")


sec_ws = sec_wb['D1BC6ACA-A78F-4688-9C6D-42D84ED']
stud_ws = stud_wb['Sheet1']

# Creates a dictionary of courseID to number of requests for that course
# Note that variables sectID and sectID2 actually hold courseID data.
courseRequests = {}
r = 2
while stud_ws.cell(row=r, column=4).value:
    if stud_ws.cell(row=r,column=5).value == "See Counselor":
        r+=1
        continue
    sectID = stud_ws.cell(row=r, column=4).value.__str__()
    sectID2 = None
    if sectID == "503000":
        sectID = "3190T1"
        sectID2 = "313753"
    elif sectID == "505000":
        sectID = "313754"
        sectID2 = "316055"
    elif sectID == "501000":
        sectID = "3190T1"
        sectID2 = "314351"
    elif sectID == "502000":
        sectID = "313753"
        sectID2 = "313754"
    elif sectID == "504000":
        sectID = "3190T1"
        sectID2 = "313754"
    elif sectID == "231904":
        sectID = "231905"
    elif sectID == "2340T1":
        sectID = "234093"
    if not sectID in set(courseRequests.keys()):
        courseRequests[sectID] = 0
    courseRequests[sectID] += 1
    if sectID2:
        if not sectID2 in set(courseRequests.keys()):
            courseRequests[sectID2] = 0
        courseRequests[sectID2] += 1
    r += 1


# Gets totals for spaces in each course from SecList workbook
courseMaxes = {}
r = 2
while sec_ws.cell(row=r, column=4).value:
    if re.search('Homeroom', sec_ws.cell(row=r, column=5).value):
        r+=1
        continue
    sectID = sec_ws.cell(row=r, column=4).value.__str__()
    if sectID == "3190T1" or sectID == "313753" or sectID == "313754" or sectID == "316055" or sectID == "314351" or sectID == "3199T1" or sectID == "3199T2" or sectID == "317860" or sectID == "319800":
        maxStuds = 28
    elif sectID == "319966" or sectID == "319967":
        maxStuds = 25
    elif sectID == "319916" or sectID == "319917" or sectID == "3199J2" or sectID == "3199J1":
        maxStuds = 23
    else:
        maxStuds = sec_ws.cell(row=r, column=9).value
    if not sectID in courseMaxes.keys():
        courseMaxes[sectID] = 0
    courseMaxes[sectID] += maxStuds
    r += 1

unfulfillable = 0
for course in courseRequests.keys():
    if not course in courseMaxes.keys():
        courseMaxes[course] = 0
    if courseRequests[course] - courseMaxes[course] >= 0:
        unfulfillable += courseRequests[course]-courseMaxes[course]

total = sum([y for y in courseRequests.values()])
print("Total course requests: " + total.__str__())
print("Unfillable course requests: " + unfulfillable.__str__())