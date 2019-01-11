from openpyxl import load_workbook, Workbook
wb=load_workbook('runs/constraint_files/SecList 8-14@1300.xlsx',read_only=True)
ws=wb['D1BC6ACA-A78F-4688-9C6D-42D84ED']
teacher_dict={}
for row in ws.rows:
    period=row[1].value
    sem=row[2].value
    crs=row[4].value
    teacher=row[5].value
    if teacher is None:
        teacher=''
    if teacher not in teacher_dict:
        teacher_dict[teacher]={}
    if period not in teacher_dict[teacher]:
        teacher_dict[teacher][period]=set()
    teacher_dict[teacher][period].add((sem,crs))
# for teacher,sched in teacher_dict.items():
#     print(teacher)
#     for period in range(1,8):
#         print('\t'+str(period))
#         if period not in sched:
#             print('\t\tNone')
#             continue
#         for i in sched[period]:
#             print('\t\t'+str(i))

wb_out = Workbook(write_only=True)
ws_out = wb_out.create_sheet()

teacher_list=list(teacher_dict.items())
teacher_list.sort()

ws_out.append(['Teacher',1,2,3,4,5,6,7])
for teacher,sched in teacher_list:
    ls=[str(teacher)]
    for period in range(1, 8):
        if period not in sched:
            ls.append('')
        else:
            courses=[]
            classes=list(sched[period])
            classes.sort(key=lambda i:str(0 if i[0]=='YR' else i[0])+i[1])
            for j in classes:
                courses.append(j[0]+': '+j[1])
            ls.append('; '.join(courses))
    ws_out.append(ls)
wb_out.save('runs/constraint_files/parsed_teacher_scheds.xlsx')
