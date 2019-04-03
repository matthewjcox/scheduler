'''
NOTE: Code assumes files are stored in folders with a file called read_params.txt, and that all files listed in
read_params.txt are in that folder as well.

Statistics to include
- Percentage of course requests fulfilled
- Number of students with empty periods (or just total number of empty periods on students' schedules)
- Number of student conflicts
- Number of teacher conflicts
- There are more but think about them later.

Figure out what Young can transfer to us.
Add mean empty seats and standard deviations per course to statistics.
'''

import re
import sys
import statistics
import tabulate
from openpyxl import Workbook
import openpyxl
from openpyxl.utils import get_column_letter

# First command line argument: file with run data to interpret
toInterpret = "../../runs/past_runs/" + sys.argv[1] + "/readable_schedule.txt" # "../../runs/past_runs/2018_12_07__18_52_41/readable_schedule.txt"  # Currently, ../../runs/perfect_schedule.txt

directory = re.search('/.*', toInterpret[::-1]).group(0)[::-1][:-1]

runParamsFile = open(directory + "/run_params.txt")
# runParams is an array of the run paramters (text files used in run). [(0) periods in a day, (1) classrooms.txt, (2) courses.txt, (3) teachers.txt, (4) students.txt, (5) sections.txt]
runParams = [line.strip() for line in runParamsFile.readlines()]

# Creates dictionary of course IDs to relevant info about courses from the courses.txt file
courseFile = open(directory +"/" + runParams[2])
# FIX THIS COMMAND SO IT BUILDS THE CORRECT DICTIONARY!!!!
# Dictionary contains courseID:[(0) course title, (1) long course title]
courses = {course[2]:course[:2] for course in [[thing.strip() for thing in line.split("|")] for line in courseFile.readlines()]}

# Opens the file with run information to analyze
file = open(toInterpret, 'r')

# Sections is a dictionary of sectionID:[(0) secionID, (1) period, (2) maxStudents, (3) courseID, (4) [teacherIDs], (5) room, (6) classSize, (7) [students], (8) [team_1], (9) [team_2], (10) [team_3], (11) [allowed_periods], (12) term].
sections = {}

numSections = int(file.readline()[:-1])
# Initiates a for-loop that runs once for every section in the schedule
for x in range(numSections):
    # Reads the data with the sectionID and stores it in "section" as a string.
    line = file.readline()[2:-1]
    # print(line)
    tempLine = line.split(": ")[1].strip()
    section = tempLine
    # Adds sectionID to the list associated with sectionID. Redundant, but useful for for-loops that loop over items in sections.values().
    # Also adds period to sections[section] as an int.
    sections[section] = [section, int(file.readline()[10])]
    # Temporarily store allowable periods (add to the end as the second to last index)
    nextLine = file.readline()
    if "Allowed periods" in nextLine:
        temp = re.search(r'[0-9][0-9, ]*', nextLine).group(0).split(",")
        nextLine = file.readline()
    else:
        temp = []
    # Reads semester and stores it in sem to add to end of list
    if "Semester" in nextLine:
        semTemp = re.search(r'year|1|2', nextLine).group(0)
        if semTemp == "year":
            sem = "0"
        else:
            sem = semTemp
        nextLine = file.readline()
    else:
        sem = None
    # Adds max students to sections[section] as an int.
    sections[section].append(int(nextLine[21:-1]))
    # Adds courseID to sections[section] as a string.
    words = file.readline().split(" ")
    if words[1] == "Course":
        sections[section].append(words[2])
    else:
        sections[section].append(words[1])
    # print(sections[section])
    # print(sections[section][-1])
    # Adds a list of teachers to section[sections] using teacher's IDs (amreid, for instance).
    tempLine = file.readline()
    if re.search(r'\([A-Za-z]+\)', tempLine.split(": ")[1].split(", ")[0]):
        sections[section].append([re.search(r'\([A-Za-z]+\)', k).group(0)[1:-1] for k in tempLine.split(": ")[1].split(", ")])
    else:
        sections[section].append([])
    # Adds room to section[sections] as a string.
    sections[section].append(file.readline()[16:-1])
    # Reads line with student data. Stored so as to be able to access multiple times.
    nextLine = file.readline()
    # Adds class size to sections[section] as an int.
    sections[section].append(int(re.search(r'\([0-9]+\)', nextLine[13:-1]).group(0)[1:-1]))
    # Adds an empty list to section[sections] if the section has no students.
    if sections[section][6] == 0:
        sections[section].append([])
    # Adds a list of students, represented by student numbers held as strings, to sections[section] if class size > 0.
    else:
        sections[section].append([re.search(r'\([0-9]+\)', curLine).group(0)[1:-1] for curLine in re.search(': .*', nextLine).group(0)[2:].split(", ")])
    # Reads the next line to check if it has teaming data.
    nextLine = file.readline()
    # If the next line has teaming data, adds teaming data to sections[section] as list of strings of section IDs/
    if re.search('Teamed 1 with:', nextLine):
        #sections[section].append([re.search('[0-9\-A-Z]+', curLine).group(0) for curLine in nextLine[15:-1].split(", ")])
        sections[section].append([curLine.split(" ")[1][:-1] for curLine in nextLine[17:-1].split(", ")])
        nextLine = file.readline()
    # If next line does not have teaming data, adds an empty list to sections[section] to represent no teaming.
    else:
        sections[section].append([])
    if re.search('Teamed 2 with:', nextLine):
        #sections[section].append([re.search(r'[0-9\-A-Z]+', curLine).group(0) for curLine in nextLine[15:-1].split(", ")])
        # print(nextLine[17:-1].split(", ")[0].split(" "))
        sections[section].append([curLine.split(" ")[1][:-1] for curLine in nextLine[17:-1].split(", ")])
        nextLine = file.readline()
    else:
        sections[section].append([])
    if re.search('Teamed 3 with:', nextLine):
        #sections[section].append([re.search(r'[0-9\-A-Z]+', curLine).group(0) for curLine in nextLine[15:-1].split(", ")])
        sections[section].append([curLine.split(" ")[1][:-1] for curLine in nextLine[17:-1].split(", ")])
        nextLine = file.readline()
    else:sections[section].append([])
    # Adds list of allowed periods.
    sections[section].append(temp)
    # Adds term
    if sem:
        sections[section].append(sem)
    # Prints sections[section] for debugging.
    # print(sections[section])

# studentScheds is a dictionary of studentID:[(0) studentID, (1) [course requests], (2) [alternates], (3) [sections], (4) [courses received]], with course and section info in IDs.
studentScheds = {}

numStuds = int(file.readline()[:-1])
# Initiates a for-loop that runs for each student in the case.
for x in range(numStuds):
    # Finds the student ID of the current student as a string.
    curStud = re.search(r'\(.*\)', file.readline()).group(0)[1:-1]
    # Adds the student number as a key in studentScheds and initiats a list as the associated value with the first element as the student ID.
    studentScheds[curStud] = [curStud, [], [], []]
    nextLine = file.readline()
    while not re.search('Alternates:', nextLine):
        words = [l.strip() for l in nextLine.split(" ")]
        #print(words)
        if words[0] == "Course":
            studentScheds[curStud][1].append(words[1])
        else:
            studentScheds[curStud][1].append(words[0])
            #print(studentScheds[curStud][1][-1])
        nextLine = file.readline()
    nextLine = file.readline()
    while not re.search('None', nextLine) and not re.search('Schedule:', nextLine):
        raise Exception("Fix so that this is compatable with schedule section of students that does or does not include the word course")
        words = nextline.split(" ")
        #print(words)
        if words[3] == "Course":
            studentScheds[curStud][2].append(words[4])
        else:
            studentScheds[curStud][2].append(words[3])
        nextLine = file.readline()
    if re.search('None', nextLine):
        file.readline()
    nextLine = file.readline()
    while not nextLine == "\n":
        #raise Exception("Fix so that this is compatable with schedule section of students that does or does not include the word course")
        #print(nextLine)
        #print(nextLine.split(" ")[1][:-1])
        studentScheds[curStud][3].append(nextLine.split(" ")[1][:-1])
        nextLine = file.readline()
    #print(studentScheds[curStud])

for student in studentScheds.values():
    #print(student)
    student.append([])
    for section in student[3]:
        #print(section)
        student[4].append(sections[section][3])
    #print(student)

# Creates a statistic file named for the file plus _stats
statFile = open(toInterpret[:-4] + "_stats.txt", "w")

theoreticalEmptySeats = (sum([sect[2] for sect in sections.values()]) - numStuds*sum([len(stud[1]) for stud in studentScheds.values()]))
statFile.write("Minimum Empty Seats: " + theoreticalEmptySeats.__str__() + "\n")
emptySeats = sum([sect[2]-sect[6] for sect in sections.values()])
statFile.write("Empty seats: " + emptySeats.__str__() + "\n")
statFile.write("Number of empty spots in student schedules: " + (emptySeats-theoreticalEmptySeats).__str__() + "\n")
#raise Exception("This number is inaccurate because it doesn't account for semester classes. Fix!")

# Make list of all students who have empty spots in schedules
# This is incorrect due to semesters. Just so you know.
studentsWithHoles = []
for student in studentScheds.values():
    if len(student[3]) < int(runParams[0]):
        studentsWithHoles.append(student[0])
statFile.write("Students with empty spaces in schedule: " + ", ".join(studentsWithHoles) + "\n")

# Calculate percent course requests fulfilled
totalRequests = 0
fulfilledRequests = 0
unrecievedCourses = {}
for student in studentScheds.values():
    # print(student)
    unfulfilledSet = {*student[1]}.difference({*student[4]})
    totalRequests += len(student[1])
    fulfilledRequests += (len(student[1]) - len(unfulfilledSet))
    for sect in unfulfilledSet:
        if not sect in unrecievedCourses.keys():
            unrecievedCourses[sect] = 0
        unrecievedCourses[sect] += 1
percentFulfilled = (fulfilledRequests/totalRequests*100).__str__() + "%"
statFile.write("Percentage of Course Requests Fulfilled: " + percentFulfilled + "\n")

# Calculate number of student conflicts (number of periods during which one student has two sections)
# Semester conflicts should count as half a conflict
'''
BELOW IS OLD CODE FOR STUDENT CONFLICTS. Does not work for data with semester classes.
raise Exception("This doesn't handle semesters or teaming. Should have 0 student conlficts on 2019_02_27__19_17_13")
studentConflicts = 0
studentsWithConflicts = []
for student in studentScheds.values():
    periods = {}
    for sect in student[3]:
        per = sections[sect][1]
        if not per in periods.keys():
            periods[per] = 0
        periods[per] += 1
    newConflicts = sum([0 if period <= 1 else 1 for period in periods.values()])
    studentConflicts += newConflicts
    if newConflicts > 0:
        studentsWithConflicts.append(student[0])
statFile.write("Number of student conflicts: " + studentConflicts.__str__() + "\n")
statFile.write("Students with conflicts: ")
if studentsWithConflicts:
    statFile.write( ", ".join(studentsWithConflicts) + "\n")
else:
    statFile.write("None\n")
'''
# Sections is a dictionary of sectionID:[(0) secionID, (1) period, (2) maxStudents, (3) courseID, (4) [teacherIDs], (5) room, (6) classSize, (7) [students], (8) [team_1], (9) [team_2], (10) [team_3], (11) [allowed_periods], (12) term].
# studentScheds is a dictionary of studentID:[(0) studentID, (1) [course requests], (2) [alternates], (3) [sections], (4) [courses received]], with course and section info in IDs.
studentConflicts = 0
studentsWithConflicts = set()
for student in studentScheds.values():
    periods = {t: {1: [], 2: []} for t in range(1,8)}
    for sect in student[3]:
        if sections[sect][12] == "0" or sections[sect][12] == "1":
            periods[sections[sect][1]][1].append(sect)
        if sections[sect][12] == "0" or sections[sect][12] == "2":
            periods[sections[sect][1]][2].append(sect)
    for t in range(1,8):
        for l in range(1,3):
            if len(periods[t][l]) > 1:
                studentConflicts += 0.5
                studentsWithConflicts.add(student[0])
statFile.write("Number of student conflicts: " + studentConflicts.__str__() + "\n")
statFile.write("Students with conflicts: ")
if studentsWithConflicts:
    statFile.write( ", ".join(studentsWithConflicts) + "\n")
else:
    statFile.write("None\n")

# Calculate number of teacher conflicts (number of periods during which one teacher has two sections)
# Semester conflicts should count as half a conflict.
'''
OLD CODE
raise Exception("This doesn't handle semesters or teaming. Should have 0 teacher conlficts on 2019_02_27__19_17_13")
teacherCons = {}
for sect in sections.values():
    for teacher in sect[4]:
        if not teacher in teacherCons.keys():
            teacherCons[teacher] = []
        teacherCons[teacher].append(sect[1])
teachersWithCons = []
teacherConflicts = 0
for teacher in teacherCons.keys():
    periods = {}
    for period in teacherCons[teacher]:
        if not period in periods.keys():
            periods[period] = 0
        periods[period] += 1
    newConflicts = sum([0 if period <= 1 else 1 for period in periods.values()])
    teacherConflicts += newConflicts
    if newConflicts > 0:
        teachersWithCons.append(teacher)
statFile.write("Number of teacher conflicts: " + teacherConflicts.__str__() + "\n")
statFile.write("Teachers with conflicts: ")
if teachersWithCons:
    statFile.write(", ".join(teachersWithCons) + "\n")
else:
    statFile.write("None\n")'''

teacherCons = {}
for sect in sections.values():
    for teacher in sect[4]:
        if not teacher in teacherCons.keys():
            teacherCons[teacher] = []
        teacherCons[teacher].append(sect[0])
teachersWithCons = set()
teacherConflicts = 0
for teacher in teacherCons.keys():
    periods = {t:{1:[], 2:[]} for t in range(1,8)}
    for sect in teacherCons[teacher]:
        if sections[sect][12] == "0" or sections[sect][12] == "1":
            periods[sections[sect][1]][1].append(sect)
        if sections[sect][12] == "0" or sections[sect][12] == "2":
            periods[sections[sect][1]][2].append(sect)
    for t in range(1,8):
        for l in periods[t].values():
            if len(l) > 1:
                if set(l).difference(set(sections[l[0]][9]+[l[0]])):
                    teacherConflicts += 1
                    teachersWithCons.add(teacher)
statFile.write("Number of teacher conflicts: " + teacherConflicts.__str__() + "\n")
statFile.write("Teachers with conflicts: ")
if teachersWithCons:
    statFile.write(", ".join(teachersWithCons) + "\n")
else:
    statFile.write("None\n")

# Check that team_1 works (calculates the number of students in some but not all of a set of teamed sections)
#print("\n")
teamMistakes = 0
checked = set()
for sect in sections.values():
    #print()
    #print(sect)
    if sect[0] in checked:
        continue
    teamed = set()
    # This is finite. Should rewrite to be recursive if teaming grows beyond 3 sections.
    for sec in sect[8]:
        teamed.add(sec)
        #print(sec)
        for s in sections[sec][8]:
            teamed.add(s)
    sharedStuds = {}
    for sec in teamed:
        for stud in sections[sec][7]:
            if not stud in sharedStuds.keys():
                sharedStuds[stud] = 0
            sharedStuds[stud] += 1
    checked = checked.union(teamed)
    for stud in sharedStuds.values():
        if stud < len(teamed):
            teamMistakes += 1
statFile.write("Number of students in some but not all of a set of team_1 sections: " + teamMistakes.__str__() + "\n")

# Calculate the number of sections with class sizes exceeding maxStudents

# Make table of courses to number of students who requested but did not get that course
unfulfilledCourses = {}
for student in studentScheds.values():
    missedSet = set(student[1]).difference(set(student[4]))
    availablePers = {1,2,3,4,5,6,7}.difference({sections[sect][1] for sect in student[3]})
    for course in student[1]:
        if not course in unfulfilledCourses.keys():
            unfulfilledCourses[course] = [courses[course][1] + " (" + course + ")"] + [0 for x in range(int(runParams[0]))]
        if course in missedSet:
            # ADD SOMETHING such that 1 is added to unfulfilledCourses[course][period].
            # i.e. must figure out available periods for student.
            for per in availablePers:
                unfulfilledCourses[course][per] += 1/len(availablePers)
#statFile.write("\n" + tabulate.tabulate([[thing[0],sum(thing[1:])] for thing in unfulfilledCourses.values()], headers=["Course", "Number of students who requested but did not get this course"], tablefmt="grid") + "\n")

# Make a table of courses to empty spaces per period
courseToEmptyPer = {}
for sect in sections.values():
    if not sect[3] in courseToEmptyPer.keys():
        courseToEmptyPer[sect[3]] = [courses[sect[3]][1] + " (" + sect[3]+ ")"] + [0 for x in range(int(runParams[0]))]
    courseToEmptyPer[sect[3]][sect[1]] += sect[2]-sect[6]

# Calculate the number of seats available each period
# EDIT TO CORRECTLY FORM TABLE OF TOTAL EMPTY SEATS AND TOTAL SEATS BY PERIOD
perToEmpty = [{"Total Empty Seats":[]}]
for x in range(1, int(runParams[0])+1):
    perToEmpty[0]["Total Empty Seats"].append(sum([cour[x] for cour in courseToEmptyPer.values()]).__str__())
perToEmpty[0]["Total Seats"] = [0 for x in range(int(runParams[0]))]
for sect in sections.values():
    perToEmpty[0]["Total Seats"][sect[1]-1] += sect[2]
statFile.write("\nTotal Seats and Empty Seats by Period\n" + tabulate.tabulate([perToEmpty], headers=["","P1","P2","P3","P4","P5","P6","P7"], tablefmt="grid") + "\n")

# Calculate the mean empty seats and standard deviation for each course
courseToEmpty = {}
for sect in sections.values():
    curCourse = sect[3]
    if not curCourse in courseToEmpty.keys():
        courseToEmpty[curCourse] = []
    courseToEmpty[curCourse].append(sect[2]-sect[6])
courseToEmptyList = [*courseToEmpty.keys()]
table1 = {"Course":[courses[course][1] + " (" + course + ")" for course in courseToEmptyList], "Mean Empty Spaces":[sum(courseToEmpty[course])/len(courseToEmpty[course]) for course in courseToEmptyList], "Standard Deviation":[statistics.stdev(courseToEmpty[course])if len(courseToEmpty[course]) > 1 else "Undefined" for course in courseToEmptyList]}
statFile.write("\n")
statFile.write("Courses to Average Number of Empty Seats and Standard Deviation\n")
statFile.write(tabulate.tabulate(table1, headers="keys", tablefmt="grid"))

toTable = {}
for course in courseToEmptyPer.keys():
    toTable[course] = [courses[course][1] + " (" + course + ")"]
    for x in range(1,int(runParams[0])+1):
        toTable[course].append(courseToEmptyPer[course][x])
        if not course in unfulfilledCourses.keys():
            unfulfilledCourses[course] = [courses[course][1] + " (" + course + ")"] + [0 for x in range(int(runParams[0]))]
        toTable[course].append(unfulfilledCourses[course][x])
    toTable[course].insert(1,sum(courseToEmptyPer[course][1:]))
    toTable[course].insert(2,sum(unfulfilledCourses[course][1:]))

statwb = Workbook()
dest_filename = "stat_spreadsheet.xlsx"
ws1 = statwb.active
ws1.title = "Empty Seats and Students Lacking Class by Period"
ws1['A1'] = "Course"
ws1['B1'] = "Total Empty Seats"
ws1['C1'] = "Total Students Missing Class"
ws1['D1'] = "P1 Empty Seats"
ws1['E1'] = "P1 Students Missing"
ws1['F1'] = "P2 Empty Seats"
ws1['G1'] = "P2 Students Missing"
ws1['H1'] = "P3 Empty Seats"
ws1['I1'] = "P3 Students Missing"
ws1['J1'] = "P4 Empty Seats"
ws1['K1'] = "P4 Students Missing"
ws1['L1'] = "P5 Empty Seats"
ws1['M1'] = "P5 Students Missing"
ws1['N1'] = "P6 Empty Seats"
ws1['O1'] = "P6 Students Missing"
ws1['P1'] = "P7 Empty Seats"
ws1['Q1'] = "P7 Students Missing"
r = 2
for course in toTable.values():
    for l in range(len(course)):
        ws1.cell(row=r, column=l+1).value = course[l]
    r += 1
ws1.freeze_panes = 'B2'
ws1.column_dimensions['A'].width = 30
ws1.cell(row=r, column=1).value = "Total"
for i in range(1,17):
    ws1.column_dimensions[get_column_letter(i+1)].width = len(ws1.cell(row=1, column=i+1).value)-2
    ws1.cell(row=r, column=i+1).value = "=SUM(" + get_column_letter(i+1) + "2:" + get_column_letter(i+1) + (r-1).__str__() + ")"
ws1.sheet_view.zoomScale = 120
statwb.save(filename="../../runs/past_runs/" + sys.argv[1] + "/" + dest_filename)

statFile.write("\n\nP1 = Period 1\tES = Empty Seats\tSM = Students Missing\tT = Total\n")
statFile.write(tabulate.tabulate(toTable.values(), headers=["Course", "T ES", "T SM", "P1 ES", "SM P1","P2 ES", "SM P2","P3 ES", "SM P3", "P4 ES", "SM P4", "P5 ES", "SM P5", "P6 ES", "SM P6", "P7 ES", "SM P7"], tablefmt="grid"))
statFile.write("\n")